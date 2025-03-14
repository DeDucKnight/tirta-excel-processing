# Form implementation generated from reading ui file 'progressExcel.ui'
#
# Created by: PyQt6 UI code generator 6.4.2
#
# WARNING: Any manual changes made to this file will be lost when pyuic6 is
# run again.  Do not edit this file unless you know what you are doing.


from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtWidgets import QFileDialog, QMessageBox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import uuid
import re

def int_to_roman(num):
        val = [
            1000, 900, 500, 400,
            100, 90, 50, 40,
            10, 9, 5, 4,
            1
        ]
        syms = [
            "M", "CM", "D", "CD",
            "C", "XC", "L", "XL",
            "X", "IX", "V", "IV",
            "I"
        ]
        roman_num = ""
        i = 0
        while num > 0:
            count = num // val[i]
            roman_num += syms[i] * count
            num %= val[i]
            i += 1
        return roman_num

def parse_numbered_item(text: str, max_prefix_len: int = 4):
    """
    Checks only the first `max_prefix_len` characters of `text` to see if it
    looks like a numbering system (e.g., "A.", "1)", "3-"). If found, returns
    [number_part, item_part]. Otherwise, returns an empty list.
    """
    text = text.strip()
    # Extract up to `max_prefix_len` characters from the start
    prefix_candidate = text[:max_prefix_len]

    try:
        # Define a simple pattern: one or more letters/digits, optionally followed by ., ), or -
        # and optional spaces. e.g., "1.", "A)", "3-", "2) "
        pattern = r'^([A-Za-z0-9]+[\.\)\-]?)\s*$'
        match = re.match(pattern, prefix_candidate.strip())
        if match:
            # We found a numbering prefix in the first few characters
            number_str = match.group(1).rstrip(".)-")  # e.g. "1", "A", "3"
            # The rest of the text is the item part
            item_part = text[len(prefix_candidate):].strip()
            return [number_str, item_part]
        else:
            # No numbering found in the first few chars
            return []
    except Exception as e:
        return []

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(724, 483)
        self.centralwidget = QtWidgets.QWidget(parent=MainWindow)
        self.centralwidget.setObjectName("centralwidget")

        # Upload Button
        self.uploadButton = QtWidgets.QPushButton(parent=self.centralwidget)
        self.uploadButton.setGeometry(QtCore.QRect(20, 10, 111, 31))
        self.uploadButton.setObjectName("uploadButton")

        # Label to Display File Name
        self.fileNameLabel = QtWidgets.QLabel(parent=self.centralwidget)
        self.fileNameLabel.setGeometry(QtCore.QRect(160, 10, 461, 31))
        self.fileNameLabel.setObjectName("fileNameLabel")

        # Sheet Selection Description Label
        self.sheetLabel = QtWidgets.QLabel(parent=self.centralwidget)
        self.sheetLabel.setGeometry(QtCore.QRect(20, 50, 200, 31))
        self.sheetLabel.setObjectName("sheetLabel")
        self.sheetLabel.setText("Pilih worksheet:")
        self.sheetLabel.setVisible(False)

        # Sheet Selection ComboBox
        self.sheetComboBox = QtWidgets.QComboBox(parent=self.centralwidget)
        self.sheetComboBox.setGeometry(QtCore.QRect(20, 80, 300, 31))
        self.sheetComboBox.setObjectName("sheetComboBox")
        self.sheetComboBox.setVisible(False)

        # Number Group Box
        self.numberGroupBox = QtWidgets.QGroupBox(parent=self.centralwidget)
        self.numberGroupBox.setGeometry(QtCore.QRect(20, 120, 681, 101))
        self.numberGroupBox.setObjectName("numberGroupBox")
        self.numYesRadioButton = QtWidgets.QRadioButton(parent=self.numberGroupBox)
        self.numYesRadioButton.setGeometry(QtCore.QRect(10, 30, 89, 20))
        self.numYesRadioButton.setObjectName("numYesRadioButton")
        self.numNoRadioButton = QtWidgets.QRadioButton(parent=self.numberGroupBox)
        self.numNoRadioButton.setGeometry(QtCore.QRect(10, 60, 91, 20))
        self.numNoRadioButton.setObjectName("numNoRadioButton")

        # Generate Button
        self.generateButton = QtWidgets.QPushButton(parent=self.centralwidget)
        self.generateButton.setGeometry(QtCore.QRect(590, 390, 111, 31))
        self.generateButton.setObjectName("generateButton")
        self.generateButton.setEnabled(False)  # Disabled until a file is chosen

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(parent=MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 724, 22))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(parent=MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actiontest1 = QtGui.QAction(parent=MainWindow)
        self.actiontest1.setObjectName("actiontest1")
        self.actiontest2 = QtGui.QAction(parent=MainWindow)
        self.actiontest2.setObjectName("actiontest2")

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        # Connect button click events to custom slots
        self.uploadButton.clicked.connect(self.on_uploadButton_clicked)
        self.generateButton.clicked.connect(self.on_generateButton_clicked)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.uploadButton.setText(_translate("MainWindow", "Pilih file"))
        self.fileNameLabel.setText(_translate("MainWindow", "Tidak ada file"))
        self.sheetLabel.setText(_translate("MainWindow", "Pilih worksheet:"))
        self.numberGroupBox.setTitle(_translate("MainWindow", "Apakah file memiliki sistem penomoran?"))
        self.numYesRadioButton.setText(_translate("MainWindow", "Ada"))
        self.numNoRadioButton.setText(_translate("MainWindow", "Tidak ada"))
        self.generateButton.setText(_translate("MainWindow", "Buat excel"))

    def on_uploadButton_clicked(self):
        options = QFileDialog.Option.DontUseNativeDialog
        file_filter = "Excel Files (*.xlsx *.xls)"  # Restrict to Excel files only
        file_name, _ = QFileDialog.getOpenFileName(None, "Select Excel File", "", file_filter, options=options)

        if file_name:
            self.fileNameLabel.setText(file_name)  # Display the selected file path
            try:
                # Store the file path for later use
                self.file_path = file_name

                # Get available sheet names using pandas ExcelFile
                xls = pd.ExcelFile(file_name)
                sheet_names = xls.sheet_names

                # Populate the combo box with the sheet names and show it along with its label
                self.sheetComboBox.clear()
                self.sheetComboBox.addItems(sheet_names)
                self.sheetComboBox.setVisible(True)
                self.sheetLabel.setVisible(True)

                # Enable the generate button since a file has been chosen
                self.generateButton.setEnabled(True)

                # Optionally, show a preview of the first sheet
                df_preview = pd.read_excel(file_name, sheet_name=sheet_names[0])
                preview = df_preview.head().to_string(index=False)
                QMessageBox.information(None, "Excel Preview", f"Sheet: {sheet_names[0]}\nFirst 5 Rows:\n{preview}")

            except Exception as e:
                QMessageBox.critical(None, "Error", f"Failed to read file:\n{str(e)}")

    def on_generateButton_clicked(self):
        # 1. Ensure a file is selected & read the user’s chosen sheet
        if not hasattr(self, "file_path"):
            QMessageBox.warning(None, "Warning", "Please upload an Excel file first.")
            return

        selected_sheet = self.sheetComboBox.currentText()

        try:
            # Read excel file
            df = pd.read_excel(self.file_path, sheet_name=selected_sheet, dtype=str)

            # If numYesRadioButton is checked, ignore the first column of the uploaded Excel
            material_column = 0
            price_column = 4
            if self.numYesRadioButton.isChecked():
                material_column = 1
                price_column = 5

            # Create a new Excel file
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Formatted Data"

            # Set column widths (tweak as needed)
            ws.column_dimensions["A"].width = 3
            ws.column_dimensions["B"].width = 7
            ws.column_dimensions["C"].width = 89
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 15
            ws.column_dimensions["F"].width = 13
            ws.column_dimensions["G"].width = 14

            start_header_row = 11
            ws.row_dimensions[start_header_row].height = 25

            # Styles
            # Fill
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            # Borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            header_border = Border(
                left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )
            
            left_cell_border = Border(
                left=Side(style='thick'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            right_cell_border = Border(
                left=Side(style='thin'),
                right=Side(style='thick'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            material_cell_border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Alignments
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_alignment = Alignment(horizontal="left", vertical="center")
            right_alignment = Alignment(horizontal="right", vertical="center")

            # Header Titles

            # Cell D2: "CV. TIRTA KUSUMA"
            ws["D2"].value = "CV. TIRTA KUSUMA"
            ws["D2"].font = Font(name="Cooper Black", size=20, bold=True)
            ws["D2"].alignment = Alignment(horizontal="center", vertical="center")

            # Cell D3: "General Contractor"
            ws["D3"].value = "General Contractor"
            ws["D3"].font = Font(name="Bodoni MT Black", size=14, bold=True)
            ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

            # Cell D4: "Jl. Manyar Rejo V / 26, Surabaya 60118. Telp.  : 031-5912879, Fax  : 031-5913274"
            ws["D4"].value = "Jl. Manyar Rejo V / 26, Surabaya 60118. Telp.  : 031-5912879, Fax  : 031-5913274"
            ws["D4"].font = Font(name="Constantia", size=10)
            ws["D4"].alignment = Alignment(horizontal="center", vertical="center")

            # Cell D5: "E-mail : tirtakusuma26@yahoo.com"
            ws["D5"].value = "E-mail : tirtakusuma26@yahoo.com"
            ws["D5"].font = Font(name="Constantia", size=10)
            ws["D5"].alignment = Alignment(horizontal="center", vertical="center")

            # Thick bottom border on row 5, columns B through G
            thick_side = Side(style='thick')
            for col in range(2, 8):  # B=2 through G=7
                cell = ws.cell(row=5, column=col)
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=thick_side
                )

            # Excel header titles
            header_titles = [
                ("B", "No"),
                ("C", "Uraian - Pekerjaan"),
                ("D", "Bobot % (I)"),
                ("E", "Presentase Pekerjaan (II)"),
                ("F", "Progress (I x II)"),
                ("G", "Kumulatif Progress %"),
            ]
            for col_letter, text in header_titles:
                cell = ws[f"{col_letter}{start_header_row}"]
                cell.value = text
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = center_alignment
                cell.font = Font(bold=True)

            # Write the data
            data_start_row = start_header_row + 1
            title_row = 1
            subsection_row = 0
            subsection_total = 0
            last_subsection_row = 0

            for i, row_data in enumerate(df.values, start=data_start_row):
                # i is the row in the *new* Excel
                # row_data is the entire row from the DataFrame

                if len(row_data) < 5:
                    continue

                # Column B in new Excel = df col 0 = Numbering
                cell_b = ws.cell(row=i, column=2)
                cell_b.border = left_cell_border
                cell_b.alignment = center_alignment
                if self.numYesRadioButton.isChecked():
                        cell_b.value = row_data[0]


                # Column C in new Excel = df col 1 = Material
                cell_c = ws.cell(row=i, column=3, value=row_data[material_column])
                cell_c.border = material_cell_border
                cell_c.alignment = left_alignment

                # Column D in new Excel = df col 4 = Bobot % (I)
                cell_d = ws.cell(row=i, column=4)
                cell_d.border = thin_border
                cell_d.alignment = right_alignment

                # Logic for price column
                price_str = str(row_data[price_column]).strip()

                if price_str.upper() == "T":
                    if self.numYesRadioButton.isChecked() is not True:
                        parse_item = parse_numbered_item(cell_c.value)
                        if parse_item:
                            cell_c.value = parse_item[1]
                            cell_b.value = parse_item[0]
                        else:
                            cell_b.value = int_to_roman(title_row)
                        title_row += 1

                    cell_c.font = Font(bold=True, italic=True)

                    # calculate subtotal
                    if last_subsection_row > 1 and subsection_total > 0:
                        cell_subsection_total_label = ws.cell(row=last_subsection_row+1, column=3)
                        cell_subsection_total_label.border = Border(
                            left=cell_subsection_total_label.border.left,
                            right=cell_subsection_total_label.border.right,
                            top=Side(style='double'),
                            bottom=Side(style='double')
                        )

                        cell_subsection_total_value = ws.cell(row=last_subsection_row+1, column=4)
                        cell_subsection_total_value.value = subsection_total
                        cell_subsection_total_value.border = Border(
                            left=cell_subsection_total_value.border.left,
                            right=cell_subsection_total_value.border.right,
                            top=Side(style='double'),
                            bottom=Side(style='double')
                        )

                        # double line border
                        cell_subtotal_e = ws.cell(row=last_subsection_row+1, column=5)
                        cell_subtotal_e.border = Border(
                            left=cell_subtotal_e.border.left,
                            right=cell_subtotal_e.border.right,
                            top=Side(style='double'),
                            bottom=Side(style='double')
                        )

                        cell_subtotal_f = ws.cell(row=last_subsection_row+1, column=6)
                        cell_subtotal_f.border = Border(
                            left=cell_subtotal_f.border.left,
                            right=cell_subtotal_f.border.right,
                            top=Side(style='double'),
                            bottom=Side(style='double')
                        )

                        cell_subtotal_g = ws.cell(row=last_subsection_row+1, column=7)
                        cell_subtotal_g.border = Border(
                            left=cell_subtotal_g.border.left,
                            right=cell_subtotal_g.border.right,
                            top=Side(style='double'),
                            bottom=Side(style='double')
                        )

                        # reset subsection row
                        subsection_row = 0
                        subsection_total = 0
                        last_subsection_row = 0
                elif price_str in ["-", "", "nan"]:
                    # It's a dash or empty cell
                    cell_d.value = ""
                else:
                    try:
                        # Attempt to parse as float
                        price_float = float(price_str.replace(".", ""))
                        cell_d.value = price_float
                        cell_b.value = subsection_row
                        subsection_row += 1
                        subsection_total += price_float
                        last_subsection_row = i

                        # if self.numYesRadioButton.isChecked() is not True:
                        #     parse_item = parse_numbered_item(cell_c.value)
                        #     if parse_item:
                        #         cell_c.value = parse_item[1]
                        #         cell_b.value = parse_item[0]
                        #     else:
                        #         cell_b.value = subsection_row
                        #     cell_b.value = subsection_row
                            
                        # If successful, do your calculation
                        # e.g., total_price = price_float * some_factor

                    except ValueError:
                        # It's not numeric, just copy it
                        cell_d.value = ""

                # Column E in new Excel
                cell_e = ws.cell(row=i, column=5, value="-")
                cell_e.border = thin_border
                cell_e.alignment = right_alignment

                # Column F in new Excel
                cell_f = ws.cell(row=i, column=6, value="-")
                cell_f.border = thin_border
                cell_f.alignment = right_alignment

                # Column G in new Excel
                cell_g = ws.cell(row=i, column=7, value="-")
                cell_g.border = right_cell_border
                cell_g.alignment = right_alignment

            # Last row
            end_row = data_start_row + len(df) - 1  # if df has X rows, last row is start + X - 1

            # Merge rows
            merge_start = end_row + 1
            merge_end = end_row + 2

            # MERGE FOR COLUMN B
            merged_cell_b = ws.cell(row=merge_start, column=2)
            merged_cell_b.value = ""
            merged_cell_b.alignment = center_alignment
            merged_cell_b.border = Border(
                left=Side(style='thick'),
                right=Side(style='thin'),
                top=Side(style='double'),
                bottom=Side(style='thick')
            )
            merged_cell_b.font = Font(bold=True)

            # MERGE FOR COLUMN C
            merged_cell_c = ws.cell(row=merge_start, column=3)
            merged_cell_c.value = "TOTAL"
            merged_cell_c.alignment = center_alignment
            merged_cell_c.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='double'),
                bottom=Side(style='thick')
            )
            merged_cell_c.font = Font(bold=True)

            # MERGE FOR COLUMN D
            merged_cell_d = ws.cell(row=merge_start, column=4)
            merged_cell_d.value = 100.00
            merged_cell_d.alignment = right_alignment
            merged_cell_d.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='double'),
                bottom=Side(style='thick')
            )

            # MERGE FOR COLUMN E
            merged_cell_e = ws.cell(row=merge_start, column=5)
            merged_cell_e.value = "-"
            merged_cell_e.alignment = right_alignment
            merged_cell_e.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='double'),
                bottom=Side(style='thick')
            )

            # MERGE FOR COLUMN F
            merged_cell_f = ws.cell(row=merge_start, column=6)
            merged_cell_f.value = "-"
            merged_cell_f.alignment = right_alignment
            merged_cell_f.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='double'),
                bottom=Side(style='thick')
            )

            # MERGE FOR COLUMN G
            merged_cell_g = ws.cell(row=merge_start, column=7)
            merged_cell_g.value = "-"
            merged_cell_g.alignment = right_alignment
            merged_cell_g.border = Border(
                left=Side(style='thin'),
                right=Side(style='thick'),
                top=Side(style='double'),
                bottom=Side(style='thick')
            )

            # 11. Save the new Excel file
            guid = uuid.uuid4()
            output_file = f"formatted_output_{guid}.xlsx"
            wb.save(output_file)

            # 11. Notify the user
            QMessageBox.information(None, "Success", f"Data formatted and saved to {output_file}")

        except Exception as e:
            QMessageBox.critical(None, "Error", f"Failed to create formatted Excel:\n{str(e)}")

    def on_headerYesRadioButton_toggled(self, checked):
        pass
        # Show or hide the header row input controls based on the radio button state
        # self.headerRowLabel.setVisible(checked)
        # self.headerRowLineEdit.setVisible(checked)
        # self.headerRowLineEdit.setEnabled(checked)


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec())

    # to do:
    # remove header
    # fix numbering system
    # subtotal per section
    # check title format with "T"
